"""Excel converter for registers, matrices, and backlogs."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from shared.converter_interface import DocumentConverter
from shared.markdown_parser import (
    extract_sections,
    parse_table_block,
    parse_test_scenarios,
    parse_user_stories,
    sanitize_sheet_name,
)
from shared.models import ConversionJob


def _style_header_row(ws, row_idx: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_column_width(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_table_sheet(wb: Workbook, title: str, rows: list[list[str]]) -> None:
    if not rows:
        return
    ws = wb.create_sheet(sanitize_sheet_name(title))
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_header_row(ws)
    _auto_column_width(ws)


def _write_records_sheet(wb: Workbook, sheet_name: str, records: list[dict[str, str]]) -> None:
    if not records:
        return
    headers = list(records[0].keys())
    ws = wb.create_sheet(sanitize_sheet_name(sheet_name))
    for c, header in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=header)
    for r, record in enumerate(records, start=2):
        for c, header in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=record.get(header, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _style_header_row(ws)
    _auto_column_width(ws)


def _convert_tables_generic(body: str, wb: Workbook) -> None:
    sections = extract_sections(body)
    used: set[str] = set()
    for heading, content in sections:
        lines = content.splitlines()
        table_lines: list[str] = []
        tables_found = 0
        for line in lines:
            if line.strip().startswith("|"):
                table_lines.append(line)
            elif table_lines:
                rows = parse_table_block(table_lines)
                if rows:
                    sheet = sanitize_sheet_name(heading if tables_found == 0 else f"{heading}_{tables_found + 1}")
                    base = sheet
                    n = 2
                    while sheet in used:
                        sheet = sanitize_sheet_name(f"{base}_{n}")
                        n += 1
                    used.add(sheet)
                    _write_table_sheet(wb, sheet, rows)
                    tables_found += 1
                table_lines = []
        if table_lines:
            rows = parse_table_block(table_lines)
            if rows:
                sheet = sanitize_sheet_name(heading)
                if sheet in used:
                    sheet = sanitize_sheet_name(f"{heading}_t")
                used.add(sheet)
                _write_table_sheet(wb, sheet, rows)


def _convert_raid_log(body: str, wb: Workbook) -> None:
    for heading in ("Risks", "Assumptions", "Issues", "Dependencies"):
        match = re.search(rf"^##\s+{heading}\s*\n(.*?)(?=^##\s+|\Z)", body, re.MULTILINE | re.DOTALL)
        if not match:
            continue
        rows = parse_table_block([line for line in match.group(1).splitlines() if line.strip().startswith("|")])
        _write_table_sheet(wb, heading, rows)


def _convert_rtm(body: str, wb: Workbook) -> None:
    for match in re.finditer(r"^##\s+(EP-\d+:.+)$", body, re.MULTILINE):
        start = match.end()
        nxt = re.search(r"^##\s+EP-\d+:", body[start:], re.MULTILINE)
        end = start + nxt.start() if nxt else len(body)
        section = body[start:end]
        rows = parse_table_block([line for line in section.splitlines() if line.strip().startswith("|")])
        if rows:
            _write_table_sheet(wb, match.group(1), rows)


class ExcelConverter(DocumentConverter):
    format_ext = ".xlsx"
    format_name = "xlsx"

    def convert(self, job: ConversionJob) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        name = Path(job.rel_path).name.lower()

        if "raid-log" in name:
            _convert_raid_log(job.body, wb)
        elif name == "rtm.md":
            _convert_rtm(job.body, wb)
        elif "epic-" in name and "stories" in name:
            _write_records_sheet(wb, "User Stories", parse_user_stories(job.body))
        elif "test-scenarios" in name:
            _write_records_sheet(wb, "Test Scenarios", parse_test_scenarios(job.body))
        elif "checklist" in name or name.endswith("-log.md"):
            _convert_tables_generic(job.body, wb)
        else:
            _convert_tables_generic(job.body, wb)

        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet("Content")
            ws.cell(row=1, column=1, value=job.meta.get("title", job.output_path.stem))
            ws.cell(row=2, column=1, value=job.body[:32000])

        wb.save(job.output_path)
